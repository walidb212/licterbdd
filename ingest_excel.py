"""
ingest_excel.py — Ingestion du dataset_decathlon.xlsx
Usage : py -3.10 ingest_excel.py --file data/dataset_decathlon.xlsx

Onglets traités :
  Reputation_Crise       → data/excel_runs/reputation_crise.jsonl
  Benchmark_Marche       → data/excel_runs/benchmark_marche.jsonl  (+ sentiment_detected via OpenRouter)
  Experience_Client_CX   → data/excel_runs/voix_client_cx.jsonl
"""
from __future__ import annotations

import argparse
import json
import os
import sys
import uuid
from datetime import datetime, timezone
from pathlib import Path

# ---------------------------------------------------------------------------
# Bootstrap : charge .env depuis le workspace
# ---------------------------------------------------------------------------
ROOT = Path(__file__).resolve().parent
sys.path.insert(0, str(ROOT))

try:
    from monitor_core import load_workspace_env
    load_workspace_env(ROOT)
except Exception:
    pass  # .env optionnel

try:
    import openpyxl
except ImportError:
    sys.exit("Dépendance manquante : pip install openpyxl")


# ---------------------------------------------------------------------------
# Colonnes inutiles à supprimer par onglet
# ---------------------------------------------------------------------------
DROP_COLS_REPUTATION = {
    "scrapingserverip", "useragentstring", "deprecatedfieldv2", "processingtimems"
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _safe_str(val: object) -> str:
    if val is None:
        return ""
    s = str(val).strip()
    return "" if s.lower() in {"none", "nan", "#n/a"} else s


def _rows_to_dicts(sheet) -> list[dict]:
    headers = [_safe_str(c.value).lower().replace(" ", "_") for c in next(sheet.iter_rows(min_row=1, max_row=1))]
    records = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        record = {headers[i]: _safe_str(v) for i, v in enumerate(row)}
        records.append(record)
    return records


def _drop_columns(records: list[dict], cols: set[str]) -> list[dict]:
    return [{k: v for k, v in r.items() if k not in cols} for r in records]


def _drop_empty_text(records: list[dict], text_col: str = "text") -> list[dict]:
    return [r for r in records if r.get(text_col, "").strip()]


def _min_word_filter(records: list[dict], text_col: str = "text", min_words: int = 5) -> list[dict]:
    return [r for r in records if len(r.get(text_col, "").split()) >= min_words]


def _write_jsonl(path: Path, records: list[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        for r in records:
            f.write(json.dumps(r, ensure_ascii=False) + "\n")
    print(f"[OK] {len(records)} lignes -> {path}")


# ---------------------------------------------------------------------------
# Enrichissement sentiment pour Benchmark_Marche
# ---------------------------------------------------------------------------

_SENTIMENT_INSTRUCTIONS = (
    "You classify brand mentions for Decathlon and Intersport. "
    "Return JSON with one 'items' array. "
    "Each item has 'row_id' (string) and 'sentiment' (one of: positive, negative, neutral). "
    "Use only evidence from the provided text. Be concise."
)

_SENTIMENT_SCHEMA = {
    "type": "object",
    "properties": {
        "items": {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {
                    "row_id": {"type": "string"},
                    "sentiment": {"type": "string", "enum": ["positive", "negative", "neutral"]},
                },
                "required": ["row_id", "sentiment"],
                "additionalProperties": False,
            },
        }
    },
    "required": ["items"],
    "additionalProperties": False,
}


def _enrich_benchmark_sentiment(records: list[dict], chunk_size: int = 8) -> list[dict]:
    api_key = os.environ.get("OPENROUTER_API_KEY", "")
    model = os.environ.get("OPENROUTER_MODEL", "openai/gpt-4o-mini")

    if not api_key:
        print("[WARN] OPENROUTER_API_KEY absent — sentiment_detected non rempli")
        return records

    from ai_batch.openrouter_client import OpenRouterChatClient
    client = OpenRouterChatClient(api_key=api_key, model=model)

    # Index records by a stable row_id
    pending = [r for r in records if not r.get("sentiment_detected", "").strip()]
    print(f"[INFO] {len(pending)} lignes sans sentiment_detected -> enrichissement OpenRouter")

    lookup = {str(i): r for i, r in enumerate(records)}
    pending_ids = [str(i) for i, r in enumerate(records) if not r.get("sentiment_detected", "").strip()]

    total = 0
    for start in range(0, len(pending_ids), chunk_size):
        chunk_ids = pending_ids[start:start + chunk_size]
        items_payload = []
        for rid in chunk_ids:
            r = lookup[rid]
            text = r.get("text") or r.get("mention_text") or r.get("content") or ""
            items_payload.append({"row_id": rid, "text": text[:800]})

        user_text = json.dumps({"items": items_payload}, ensure_ascii=False)
        try:
            payload = client.create_structured_response(
                instructions=_SENTIMENT_INSTRUCTIONS,
                user_text=user_text,
                schema_name="benchmark_sentiment_batch",
                schema=_SENTIMENT_SCHEMA,
            )
            raw = client.extract_output_text(payload)
            parsed = json.loads(raw) if raw else {"items": []}
            for item in parsed.get("items", []):
                rid = str(item.get("row_id", ""))
                sentiment = item.get("sentiment", "neutral")
                if rid in lookup:
                    lookup[rid]["sentiment_detected"] = sentiment
                    total += 1
        except Exception as exc:
            print(f"[WARN] Chunk {start//chunk_size + 1} failed: {exc}")

    print(f"[OK] {total} sentiments remplis")
    return list(lookup.values())


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(description="Ingest dataset_decathlon.xlsx")
    parser.add_argument("--file", default="data/dataset_decathlon.xlsx", help="Chemin vers le fichier Excel")
    parser.add_argument("--out-dir", default="data/excel_runs", help="Dossier de sortie")
    parser.add_argument("--skip-sentiment", action="store_true", help="Ne pas appeler OpenRouter pour Benchmark_Marche")
    args = parser.parse_args()

    xlsx_path = Path(args.file)
    if not xlsx_path.exists():
        sys.exit(f"Fichier introuvable : {xlsx_path}")

    out_dir = Path(args.out_dir)
    print(f"[INFO] Lecture de {xlsx_path} …")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    run_id = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%S%fZ") + "_" + uuid.uuid4().hex[:6]

    # ---- Onglet 1 : Reputation_Crise ----
    if "Reputation_Crise" in wb.sheetnames:
        ws = wb["Reputation_Crise"]
        rows = _rows_to_dicts(ws)
        rows = _drop_columns(rows, DROP_COLS_REPUTATION)
        rows = _drop_empty_text(rows, "text")
        rows = _min_word_filter(rows, "text", 5)
        for r in rows:
            r["_sheet"] = "reputation_crise"
            r["_run_id"] = run_id
        _write_jsonl(out_dir / "reputation_crise.jsonl", rows)
    else:
        print("[WARN] Onglet 'Reputation_Crise' introuvable")

    # ---- Onglet 2 : Benchmark_Marche ----
    bench_sheet = next((s for s in wb.sheetnames if "benchmark" in s.lower()), None)
    if bench_sheet:
        ws = wb[bench_sheet]
        rows = _rows_to_dicts(ws)
        rows = _drop_empty_text(rows, "text")
        rows = _min_word_filter(rows, "text", 5)
        for r in rows:
            r["_sheet"] = "benchmark_marche"
            r["_run_id"] = run_id
            if "sentiment_detected" not in r:
                r["sentiment_detected"] = ""
        if not args.skip_sentiment:
            rows = _enrich_benchmark_sentiment(rows)
        _write_jsonl(out_dir / "benchmark_marche.jsonl", rows)
    else:
        print("[WARN] Onglet 'Benchmark_Marche' introuvable")

    # ---- Onglet 3 : Experience_Client_CX ----
    cx_sheet = next((s for s in wb.sheetnames if "experience" in s.lower() or "voix" in s.lower() or "cx" in s.lower()), None)
    if cx_sheet:
        ws = wb[cx_sheet]
        rows = _rows_to_dicts(ws)
        rows = _drop_empty_text(rows, "text")
        for r in rows:
            r["_sheet"] = "voix_client_cx"
            r["_run_id"] = run_id
        _write_jsonl(out_dir / "voix_client_cx.jsonl", rows)
    else:
        print("[WARN] Onglet 'Experience_Client_CX' introuvable")

    wb.close()
    print(f"\n[DONE] run_id={run_id} -> {out_dir}/")


if __name__ == "__main__":
    main()
